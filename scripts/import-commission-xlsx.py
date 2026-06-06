#!/usr/bin/env python3
"""Dry-run Kari commission spreadsheet imports.

Parses historical weekly Excel sheets and prints a JSON report by default. This is intentionally
read-only: it does not write Household Hub data. Use the report to review totals, suspicious dates,
calculation mismatches, duplicate candidates, and tip-type breakdowns before importing.
"""
from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

HEADER_ALIASES = {
    'date': 'date',
    'client': 'clientName',
    'product/service': 'serviceName',
    'product service': 'serviceName',
    'service': 'serviceName',
    'revenue': 'revenue',
    'commission rate': 'commissionRate',
    'commission amount': 'commissionAmount',
    'deductions': 'deductions',
    'payout': 'payout',
    'separate tips': 'tipAmount',
    'type of tip': 'tipType',
}

NUMERIC_FIELDS = {'revenue', 'commissionRate', 'commissionAmount', 'deductions', 'payout', 'tipAmount'}
REQUIRED_FIELDS = {'date', 'revenue', 'commissionRate'}


def clean_header(value: Any) -> str:
    return re.sub(r'\s+', ' ', str(value or '').strip().lower())


def workbook_year(path: Path) -> int | None:
    match = re.search(r'(20\d{2})', path.name)
    return int(match.group(1)) if match else None


def weekday_name(iso_date: str | None) -> str:
    if not iso_date:
        return ''
    try:
        return date.fromisoformat(iso_date).strftime('%A')
    except ValueError:
        return ''


def enrich_date_suggestions(rows: list[dict[str, Any]], expected_year: int | None) -> None:
    """Infer suspicious wrong-year dates from nearby dated rows in the same weekly sheet.

    The sheet tab is a week-start label, not the service date. If a typo-year row is surrounded
    by valid dates in the same sheet, prefer that local row context over blindly replacing the
    year or using the sheet title.
    """
    if not expected_year:
        return
    by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_sheet[row.get('sourceSheet', '')].append(row)

    for sheet_rows in by_sheet.values():
        sheet_rows.sort(key=lambda row: row.get('sourceRow') or 0)
        valid = [row for row in sheet_rows if row.get('date') and row['date'].startswith(str(expected_year))]
        for row in sheet_rows:
            problems = row.get('problems') or []
            if 'date_outside_workbook_year' not in problems:
                continue
            row_num = row.get('sourceRow') or 0
            previous = next((candidate for candidate in reversed(valid) if (candidate.get('sourceRow') or 0) < row_num), None)
            following = next((candidate for candidate in valid if (candidate.get('sourceRow') or 0) > row_num), None)
            fallback = f'{expected_year}{row["date"][4:]}'
            suggested = fallback
            method = 'replace_year_preserve_month_day'
            if previous and following and previous.get('date') == following.get('date'):
                suggested = previous['date']
                method = 'neighboring_rows_same_date'
            elif previous and following:
                prev_date = date.fromisoformat(previous['date'])
                next_date = date.fromisoformat(following['date'])
                fallback_date = date.fromisoformat(fallback)
                if prev_date <= fallback_date <= next_date:
                    suggested = fallback
                    method = 'between_neighboring_dates'
                else:
                    # Prefer the closer surrounding work date inside the same sheet when the
                    # month/day from the typo-year date falls outside the local row sequence.
                    suggested = previous['date'] if abs((fallback_date - prev_date).days) <= abs((next_date - fallback_date).days) else following['date']
                    method = 'nearest_neighboring_date'
            elif previous:
                suggested = previous['date']
                method = 'previous_row_date'
            elif following:
                suggested = following['date']
                method = 'following_row_date'

            row['suggestedDate'] = suggested
            row['suggestionMethod'] = method
            row['dateEvidence'] = {
                'originalDate': row.get('date'),
                'originalWeekday': weekday_name(row.get('date')),
                'suggestedDate': suggested,
                'suggestedWeekday': weekday_name(suggested),
                'previousDate': previous.get('date') if previous else '',
                'previousWeekday': weekday_name(previous.get('date')) if previous else '',
                'previousRow': previous.get('sourceRow') if previous else None,
                'nextDate': following.get('date') if following else '',
                'nextWeekday': weekday_name(following.get('date')) if following else '',
                'nextRow': following.get('sourceRow') if following else None,
            }


def to_float(value: Any) -> float:
    if value in (None, ''):
        return 0.0
    if isinstance(value, str):
        value = value.replace('$', '').replace(',', '').strip()
        if value.endswith('%'):
            try:
                return float(value[:-1]) / 100
            except ValueError:
                return math.nan
    try:
        number = float(value)
        return round(number, 2) if abs(number) >= 1 else number
    except (TypeError, ValueError):
        return math.nan


def excel_date(value: Any) -> str | None:
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and value > 20000:
        # Excel serial date, using the 1899-12-30 convention used by openpyxl.
        return (date(1899, 12, 30) + timedelta(days=int(value))).isoformat()
    if isinstance(value, str):
        text = value.strip()
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%m-%d-%Y', '%m-%d-%y'):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                pass
    return None


def find_header_row(sheet) -> tuple[int, dict[int, str]] | None:
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10), values_only=False):
        mapped = {}
        for cell in row:
            header = HEADER_ALIASES.get(clean_header(cell.value))
            if header:
                mapped[cell.column] = header
        if {'date', 'clientName', 'serviceName', 'revenue'}.issubset(set(mapped.values())):
            return row[0].row, mapped
    return None


def parse_workbook(path: Path, include_rows: bool = False) -> dict[str, Any]:
    expected_year = workbook_year(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    rows = []
    problems = []
    totals = defaultdict(float)
    tips_by_type = defaultdict(float)
    service_counter = Counter()
    duplicate_keys = Counter()
    unusual_rates = Counter()

    for sheet in wb.worksheets:
        if 'template' in sheet.title.strip().lower():
            continue
        header = find_header_row(sheet)
        if not header:
            continue
        header_row, mapping = header
        for row_number, cells in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            raw = {field: cells[col - 1] if col - 1 < len(cells) else None for col, field in mapping.items()}
            if not any(value not in (None, '') for value in raw.values()):
                continue
            first_text = str(raw.get('date') or '').strip().lower()
            if first_text in {'total', 'totals'}:
                continue

            parsed_date = excel_date(raw.get('date'))
            if not parsed_date:
                continue

            # Real work rows always have a date plus at least service/client or dollars.
            # Some source sheets contain pre-formatted blank rows whose formula cells
            # resolve to 0 in data_only mode; do not count those as import candidates.
            has_identity = any(str(raw.get(field) or '').strip() for field in ('clientName', 'serviceName'))
            has_money = any(to_float(raw.get(field)) not in (0, 0.0) for field in NUMERIC_FIELDS)
            if not has_identity and not has_money:
                continue

            entry = {
                'date': parsed_date,
                'clientName': str(raw.get('clientName') or '').strip(),
                'serviceName': str(raw.get('serviceName') or '').strip(),
                'sourceWorkbook': path.name,
                'sourceSheet': sheet.title,
                'sourceRow': row_number,
            }
            for field in NUMERIC_FIELDS:
                entry[field] = to_float(raw.get(field))
            entry['tipType'] = str(raw.get('tipType') or '').strip()

            row_problems = []
            if not parsed_date:
                row_problems.append('invalid_date')
            elif expected_year and int(parsed_date[:4]) != expected_year:
                original_date = parsed_date
                corrected_date = f'{expected_year}{parsed_date[4:]}'
                entry['date'] = corrected_date
                entry['originalDate'] = original_date
                entry['autoFixed'] = ['date_year']
                entry['dateEvidence'] = {
                    'originalDate': original_date,
                    'originalWeekday': weekday_name(original_date),
                    'correctedDate': corrected_date,
                    'correctedWeekday': weekday_name(corrected_date),
                    'reason': 'year replaced with workbook year; month/day preserved',
                }
            for field in REQUIRED_FIELDS:
                if field == 'date':
                    continue
                if not math.isfinite(entry.get(field, math.nan)):
                    row_problems.append(f'invalid_{field}')
            if not entry['clientName'] and not entry['serviceName']:
                row_problems.append('missing_client_and_service')

            calc_commission = round((entry['revenue'] or 0) * (entry['commissionRate'] or 0), 2)
            if math.isfinite(entry['commissionAmount']) and abs(calc_commission - entry['commissionAmount']) > 0.02:
                row_problems.append('commission_mismatch')
                entry['calculatedCommissionAmount'] = calc_commission
            calc_payout = round((entry['commissionAmount'] or 0) - (entry.get('deductions') or 0), 2)
            if math.isfinite(entry['payout']) and abs(calc_payout - entry['payout']) > 0.02:
                row_problems.append('payout_mismatch')
                entry['calculatedPayout'] = calc_payout
            rate = entry['commissionRate']
            if math.isfinite(rate) and (rate <= 0 or rate >= 1):
                unusual_rates[str(rate)] += 1

            key = (entry['date'], entry['clientName'].lower(), entry['serviceName'].lower(), entry['revenue'])
            duplicate_keys[key] += 1
            for field in ('revenue', 'commissionAmount', 'deductions', 'payout', 'tipAmount'):
                if math.isfinite(entry[field]):
                    totals[field] += entry[field]
            if entry['tipAmount'] and math.isfinite(entry['tipAmount']):
                tips_by_type[entry['tipType'] or 'Other'] += entry['tipAmount']
            if entry['serviceName']:
                service_counter[entry['serviceName']] += 1

            if row_problems:
                problems.append({**entry, 'problems': row_problems})
            rows.append({
                **entry,
                'problems': row_problems,
                'needsReview': bool(row_problems),
                'reviewReason': ','.join(row_problems),
            })

    duplicates = [
        {'date': key[0], 'clientName': key[1], 'serviceName': key[2], 'revenue': key[3], 'count': count}
        for key, count in duplicate_keys.items()
        if count > 1 and key[0]
    ]
    report = {
        'workbook': path.name,
        'expectedYear': expected_year,
        'sheetCount': len(wb.worksheets),
        'rowCount': len(rows),
        'totals': {key: round(value, 2) for key, value in totals.items()},
        'tipsByType': {key: round(value, 2) for key, value in sorted(tips_by_type.items())},
        'problemCount': len(problems),
        'problems': problems[:50],
        'duplicateCandidates': duplicates[:50],
        'duplicateCandidateCount': len(duplicates),
        'unusualRates': dict(unusual_rates),
        'topServices': service_counter.most_common(15),
    }
    if include_rows:
        report['rows'] = rows
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description='Dry-run commission spreadsheet import.')
    parser.add_argument('workbooks', nargs='+', help='Excel .xlsx files to inspect')
    parser.add_argument('--pretty', action='store_true', help='Pretty-print JSON')
    parser.add_argument('--include-rows', action='store_true', help='Include normalized import candidate rows in the JSON report')
    args = parser.parse_args()

    reports = [parse_workbook(Path(path), include_rows=args.include_rows) for path in args.workbooks]
    aggregate = {
        'workbookCount': len(reports),
        'rowCount': sum(report['rowCount'] for report in reports),
        'problemCount': sum(report['problemCount'] for report in reports),
        'duplicateCandidateCount': sum(report['duplicateCandidateCount'] for report in reports),
        'totals': {},
    }
    for report in reports:
      for key, value in report['totals'].items():
        aggregate['totals'][key] = round(aggregate['totals'].get(key, 0) + value, 2)

    print(json.dumps({'aggregate': aggregate, 'workbooks': reports}, indent=2 if args.pretty else None, default=str))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
